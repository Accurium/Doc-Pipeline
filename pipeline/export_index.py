"""
export_index.py
===============
Queries the DuckDB database and writes three Excel files to the OneDrive
exports folder (2. File processing (Input - Output)/exports/):

  1. Index_[timestamp].xlsx
       Pivot: APIR code × document type. Each cell contains the renamed filename.
       Includes sha256 column for traceability.

  2. Masterlist_[timestamp].xlsx
       One row per (APIR code × document type). Shows fund name, status,
       file path, sha256, confidence, date.

  3. Tracker_[timestamp].xlsx
       Two sheets:
         - 'Tracker'    : one row per APIR code, columns per doc type.
                          C = complete, NA = not available, X = missing.
                          Row turns green when all mandatory types are C or NA.
         - 'Not Available': the full not_available table for reference.

USAGE
-----
  python pipeline\export_index.py

  # Export to a different folder (override)
  python pipeline\export_index.py --output "C:\some\other\folder"
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import DB_PATH, EXPORTS_DIR, MANDATORY_DOC_TYPES, ALL_DOC_TYPES

try:
    import duckdb
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "duckdb"])
    import duckdb

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FONT     = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL     = PatternFill("solid", fgColor="2E4057")
BODY_FONT       = Font(name="Arial", size=10)
COMPLETE_FILL   = PatternFill("solid", fgColor="C6EFCE")   # green
MISSING_FILL    = PatternFill("solid", fgColor="FFC7CE")   # red
NA_FILL         = PatternFill("solid", fgColor="FFEB9C")   # amber
ROW_FILL_ALT    = PatternFill("solid", fgColor="F5F5F5")
THIN_BORDER     = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header_row(ws, row_num: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER


def style_body_cell(cell, alt_row: bool = False):
    cell.font      = BODY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = THIN_BORDER
    if alt_row:
        cell.fill = ROW_FILL_ALT


def autofit(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 2, min_width), max_width
        )


# ── Queries ───────────────────────────────────────────────────────────────────

FILES_QUERY = """
SELECT
    f.apir_code,
    r.official_name,
    r.mstar_io_name,
    f.doc_type,
    f.renamed_name,
    f.doc_date,
    f.confidence,
    f.sha256,
    f.renamed_path,
    f.ingested_at
FROM files f
LEFT JOIN apir_reference r ON r.apir_code = f.apir_code
WHERE f.status = 'renamed'
ORDER BY f.apir_code, f.doc_type
"""

NOT_AVAILABLE_QUERY = """
SELECT apir_code, doc_type, reason, updated_at, updated_by
FROM not_available
ORDER BY apir_code, doc_type
"""

APIR_REF_QUERY = """
SELECT apir_code, official_name, mstar_io_name
FROM apir_reference
ORDER BY apir_code
"""


# ── Index workbook ────────────────────────────────────────────────────────────

def build_index(con, output_path: Path):
    rows  = con.execute(FILES_QUERY).fetchall()
    cols  = [d[0] for d in con.description]

    # Build pivot: apir_code → doc_type → list of (renamed_name, sha256)
    doc_types = sorted(ALL_DOC_TYPES)
    pivot: dict[str, dict] = {}
    apir_names: dict[str, str] = {}

    for row in rows:
        r = dict(zip(cols, row))
        apir = r["apir_code"] or "_unknown"
        dt   = r["doc_type"]  or "_unknown"
        pivot.setdefault(apir, {}).setdefault(dt, []).append(r)
        apir_names[apir] = r.get("mstar_io_name") or r.get("official_name") or ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Index"
    ws.freeze_panes = "C2"

    # Header
    headers = ["APIR / Ticker", "Fund name"] + doc_types
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    # Data rows
    all_apirs = sorted(pivot.keys())
    for row_idx, apir in enumerate(all_apirs, 2):
        alt = row_idx % 2 == 0
        ws.cell(row=row_idx, column=1, value=apir)
        style_body_cell(ws.cell(row=row_idx, column=1), alt)
        ws.cell(row=row_idx, column=2, value=apir_names.get(apir, ""))
        style_body_cell(ws.cell(row=row_idx, column=2), alt)

        for col_idx, dt in enumerate(doc_types, 3):
            entries = pivot[apir].get(dt, [])
            if entries:
                # Show renamed filename(s), pipe-separated if multiple
                cell_val = " | ".join(e["renamed_name"] for e in entries if e["renamed_name"])
            else:
                cell_val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            style_body_cell(cell, alt)

    autofit(ws)
    wb.save(output_path)
    print(f"  Index written:      {output_path.name}  ({len(all_apirs)} APIR codes)")


# ── Masterlist workbook ───────────────────────────────────────────────────────

def build_masterlist(con, output_path: Path):
    rows = con.execute(FILES_QUERY).fetchall()
    cols = [d[0] for d in con.description]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Masterlist"
    ws.freeze_panes = "A2"

    headers = [
        "APIR / Ticker", "Fund name (Mstar)", "Document type",
        "Document date", "Confidence", "Renamed filename",
        "SHA-256", "Renamed path", "Ingested at",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 24

    for row_idx, row in enumerate(rows, 2):
        r   = dict(zip(cols, row))
        alt = row_idx % 2 == 0
        values = [
            r.get("apir_code"),
            r.get("mstar_io_name") or r.get("official_name"),
            r.get("doc_type"),
            r.get("doc_date"),
            r.get("confidence"),
            r.get("renamed_name"),
            r.get("sha256"),
            r.get("renamed_path"),
            str(r.get("ingested_at") or ""),
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            style_body_cell(cell, alt)

    autofit(ws)
    wb.save(output_path)
    print(f"  Masterlist written: {output_path.name}  ({len(rows)} rows)")


# ── Tracker workbook ──────────────────────────────────────────────────────────

def build_tracker(con, output_path: Path):
    # Load all renamed files
    renamed_rows = con.execute(FILES_QUERY).fetchall()
    renamed_cols = [d[0] for d in con.description]
    renamed_set: dict[tuple, str] = {}   # (apir, doc_type) → renamed_name
    for row in renamed_rows:
        r = dict(zip(renamed_cols, row))
        renamed_set[(r["apir_code"], r["doc_type"])] = r["renamed_name"]

    # Load not_available
    na_rows = con.execute(NOT_AVAILABLE_QUERY).fetchall()
    na_set: set[tuple] = {(r[0], r[1]) for r in na_rows}

    # Load all APIR codes from reference
    apir_rows = con.execute(APIR_REF_QUERY).fetchall()

    doc_types = sorted(ALL_DOC_TYPES)
    mandatory = MANDATORY_DOC_TYPES

    wb = openpyxl.Workbook()

    # ── Sheet 1: Tracker ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tracker"
    ws.freeze_panes = "C2"

    headers = ["APIR / Ticker", "Fund name"] + doc_types + ["Complete?"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    for row_idx, apir_row in enumerate(apir_rows, 2):
        apir      = apir_row[0]
        fund_name = apir_row[2] or apir_row[1] or ""
        alt       = row_idx % 2 == 0

        ws.cell(row=row_idx, column=1, value=apir)
        style_body_cell(ws.cell(row=row_idx, column=1), alt)
        ws.cell(row=row_idx, column=2, value=fund_name)
        style_body_cell(ws.cell(row=row_idx, column=2), alt)

        mandatory_statuses = []

        for col_idx, dt in enumerate(doc_types, 3):
            key = (apir, dt)
            if key in renamed_set:
                status = "C"
                fill   = COMPLETE_FILL
            elif key in na_set:
                status = "NA"
                fill   = NA_FILL
            else:
                status = "X"
                fill   = MISSING_FILL if dt in mandatory else None

            cell = ws.cell(row=row_idx, column=col_idx, value=status)
            cell.font      = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill

            if dt in mandatory:
                mandatory_statuses.append(status)

        # Complete? column
        is_complete = all(s in ("C", "NA") for s in mandatory_statuses)
        complete_cell = ws.cell(row=row_idx, column=len(headers),
                                value="Yes" if is_complete else "No")
        complete_cell.font      = Font(name="Arial", size=10, bold=is_complete)
        complete_cell.fill      = COMPLETE_FILL if is_complete else PatternFill()
        complete_cell.alignment = Alignment(horizontal="center", vertical="center")
        complete_cell.border    = THIN_BORDER

    autofit(ws)

    # ── Sheet 2: Not Available ────────────────────────────────────────────
    ws2 = wb.create_sheet("Not Available")
    na_headers = ["APIR / Ticker", "Document type", "Reason", "Updated at", "Updated by"]
    for col_idx, h in enumerate(na_headers, 1):
        ws2.cell(row=1, column=col_idx, value=h)
    style_header_row(ws2, 1, len(na_headers))

    for row_idx, r in enumerate(na_rows, 2):
        alt = row_idx % 2 == 0
        for col_idx, v in enumerate(r, 1):
            style_body_cell(ws2.cell(row=row_idx, column=col_idx, value=str(v) if v else ""), alt)

    autofit(ws2)

    wb.save(output_path)
    complete_count = sum(
        1 for ar in apir_rows
        if all(
            ((ar[0], dt) in renamed_set or (ar[0], dt) in na_set)
            for dt in mandatory
        )
    )
    print(f"  Tracker written:    {output_path.name}  "
          f"({complete_count}/{len(apir_rows)} APIR codes complete)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export index, masterlist, tracker from DuckDB.")
    parser.add_argument("--output", default=None,
                        help="Override output folder (default: OneDrive exports dir from config).")
    args = parser.parse_args()

    out_dir = Path(args.output) if args.output else EXPORTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now(ZoneInfo("Australia/Sydney")).strftime("%Y_%m_%d_%H_%M_%S")

    con = duckdb.connect(str(DB_PATH), read_only=True)

    print(f"Writing exports to: {out_dir}")
    build_index(     con, out_dir / f"Index_{ts}.xlsx")
    build_masterlist(con, out_dir / f"Masterlist_{ts}.xlsx")
    build_tracker(   con, out_dir / f"Tracker_{ts}.xlsx")

    con.close()
    print("Done.")


if __name__ == "__main__":
    main()
