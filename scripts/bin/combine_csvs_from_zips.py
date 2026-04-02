#!/usr/bin/env python3
"""
combine_csv_zips.py
===================
Purpose
-------
This script is designed to support the APL Investment Option Assessment pipeline.
AI agents in the pipeline produce output batches delivered as zip files, each
containing CSV files with standardised naming conventions (e.g. Rename_log_*.csv,
Rename_list_*.csv, Exceptions_list_*.csv). Over time, multiple batch runs accumulate
across multiple zip files.

This script:
  1. Recursively scans a root folder for all zip files.
  2. Inside each zip, identifies CSV files matching the three known naming patterns.
  3. Reads all matching CSVs and groups their rows by file type.
  4. Deduplicates rows across files (ignoring which source file a row came from).
  5. Writes one combined output Excel (.xlsx) file per file type, with a Source_File
     column added so the origin of each row remains traceable.
  6. Produces an APRA-grade defensibility log in both CSV and JSON formats,
     recording inputs used, run timestamp, outputs produced, exceptions encountered,
     and overall run status. Log files are named csv_combine_log_[timestamp].

Output files are written to:
  ~/Documents/combined_output/
    Combined_Rename_log_[YYYYMMDD].xlsx
    Combined_Rename_list_[YYYYMMDD].xlsx
    Combined_Exceptions_list_[YYYYMMDD].xlsx
    csv_combine_log_[YYYYMMDD].csv
    csv_combine_log_[YYYYMMDD].json

Defensibility Log
-----------------
The log is designed to meet APRA-grade audit requirements and records:
  - Input files used: each zip file processed, with its full path and file size
  - Date and time the script was run (ISO 8601 format)
  - Outputs produced: each combined Excel file written, with row counts and file size
  - Exceptions and errors: corrupt zips, unreadable CSVs, schema mismatches,
    duplicate rows removed, and any unexpected errors with full detail
  - Status: COMPLETED, COMPLETED_WITH_WARNINGS, or FAILED

Configuration
-------------
Set ROOT_FOLDER (line ~60) to the folder containing your zip files before running.
All subfolders are searched recursively.

Usage
-----
  python combine_csv_zips.py

Requirements
------------
  Python 3.6+. Requires openpyxl for Excel output:
    pip install openpyxl
"""

import csv
import json
import re
import traceback
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile, BadZipFile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ==============================
# DEFINE YOUR INPUT FOLDER HERE
# Change this path to point to the folder containing your zip files.
# The script will search this folder and all subfolders recursively.
# ==============================
ROOT_FOLDER = Path(r"C:\Development\ETSL Renaming Prompt\output")

# Output folder where combined Excel files and the defensibility log will be written.
# Created automatically if it does not already exist.
OUTPUT_FOLDER = Path(r"C:\Development\ETSL Renaming Prompt\output")

# -------------------------------------------------------------------
# FILE MATCHING PATTERNS
# Each entry maps a group name to a regex pattern.
# The script uses these to identify which CSV files inside a zip
# belong to which output group.
#   Rename_log      — logs of renaming actions taken by the AI agents
#   Rename_list     — lists of items to be renamed
#   Exceptions_list — items that could not be processed or were flagged
# -------------------------------------------------------------------
PATTERNS = {
    "Rename_log":      re.compile(r"^.*Rename_log_.*\.csv$",      re.IGNORECASE),
    "Rename_list":     re.compile(r"^.*Rename_list_.*\.csv$",     re.IGNORECASE),
    "Exceptions_list": re.compile(r"^.*Exceptions_list_.*\.csv$", re.IGNORECASE),
}

# Maps each group name to its combined output filename base.
# A date suffix (YYYYMMDD) is appended at write time: e.g. Combined_Rename_log_20260324.xlsx
OUTPUT_FILE_BASES = {
    "Rename_log": "Combined_Rename_log",
    "Rename_list": "Combined_Rename_list",
    "Exceptions_list": "Combined_Exceptions_list",
}


# -------------------------------------------------------------------
# EXCEL HEADER STYLE
# Applied to the first row of each combined Excel output sheet.
# -------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_header_style(ws):
    """
    Applies bold white text on a dark blue background to the first (header) row
    of a worksheet, and sets a sensible default column width.
    """
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Auto-size columns to a reasonable width based on header label length.
    for col in ws.columns:
        header_val = col[0].value or ""
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(15, min(len(str(header_val)) + 4, 50))


# -------------------------------------------------------------------
# DEFENSIBILITY LOG STRUCTURE
# Initialised once per run. All functions append to this shared record.
# -------------------------------------------------------------------
def init_run_log(run_timestamp: str) -> dict:
    """
    Initialises the defensibility log structure for a single script run.

    The log captures all information required to reconstruct and audit
    the actions taken during this run. It is written in full at the end
    of the run in both CSV and JSON formats.

    Fields
    ------
    run_id          : Unique identifier for this run — matches the log filename timestamp.
    run_timestamp   : ISO 8601 datetime when the script started (UTC).
    script_version  : Version label for this script — increment when making changes.
    root_folder     : Absolute path of the input folder scanned.
    output_folder   : Absolute path of the output folder written to.
    status          : COMPLETED | COMPLETED_WITH_WARNINGS | FAILED
    status_detail   : Plain-language summary of the final status.
    inputs          : List of input file records (one per zip file found).
    outputs         : List of output file records (one per combined Excel file written).
    exceptions      : List of exception/error records encountered during the run.
    summary         : Aggregate counts for audit review.
    """
    return {
        "run_id": run_timestamp,
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "script_version": "3.0",
        "root_folder": str(ROOT_FOLDER.resolve()),
        "output_folder": str(OUTPUT_FOLDER.resolve()),
        "status": "COMPLETED",
        "status_detail": "",
        "inputs": [],
        "outputs": [],
        "exceptions": [],
        "summary": {
            "zip_files_found": 0,
            "zip_files_processed": 0,
            "zip_files_skipped": 0,
            "csv_files_matched": 0,
            "total_rows_read": 0,
            "total_duplicate_rows_removed": 0,
            "total_rows_written": 0,
            "output_files_written": 0,
            "exception_count": 0,
        },
    }


def log_input(run_log: dict, zip_path: Path, csv_files_found: list, rows_read: int):
    """
    Records a processed zip file in the inputs section of the defensibility log.
    """
    try:
        file_size_bytes = zip_path.stat().st_size
    except OSError:
        file_size_bytes = None

    run_log["inputs"].append({
        "file_path": str(zip_path.resolve()),
        "file_name": zip_path.name,
        "file_size_bytes": file_size_bytes,
        "matching_csv_files": csv_files_found,
        "rows_read": rows_read,
    })


def log_output(run_log: dict, output_path: Path, group: str, rows_written: int):
    """
    Records a written output file in the outputs section of the defensibility log.
    """
    try:
        file_size_bytes = output_path.stat().st_size
    except OSError:
        file_size_bytes = None

    run_log["outputs"].append({
        "file_path": str(output_path.resolve()),
        "file_name": output_path.name,
        "group": group,
        "rows_written": rows_written,
        "file_size_bytes": file_size_bytes,
    })
    run_log["summary"]["output_files_written"] += 1
    run_log["summary"]["total_rows_written"] += rows_written


def log_exception(run_log: dict, error_type: str, detail: str,
                  source_file: str = "", severity: str = "WARNING"):
    """
    Records an exception or error in the exceptions section of the defensibility log.

    Severity levels
    ---------------
    WARNING : Non-fatal — processing continued. Logged for audit visibility.
    ERROR   : A file or group could not be processed. May affect output completeness.
    CRITICAL: The run could not complete. status will be set to FAILED.
    """
    run_log["exceptions"].append({
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "severity": severity,
        "error_type": error_type,
        "source_file": source_file,
        "detail": detail,
    })
    run_log["summary"]["exception_count"] += 1

    if severity == "CRITICAL":
        run_log["status"] = "FAILED"
    elif severity in ("WARNING", "ERROR") and run_log["status"] == "COMPLETED":
        run_log["status"] = "COMPLETED_WITH_WARNINGS"


def write_defensibility_log(run_log: dict, run_timestamp: str):
    """
    Writes the defensibility log in both CSV and JSON formats.

    CSV format: flat, sectioned layout readable in Excel.
    JSON format: full nested structure for machine ingestion.

    Both files are named csv_combine_log_[YYYYMMDD].
    """
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    log_stem = f"csv_combine_log_{run_timestamp}"

    # ------------------------------------------------------------------
    # JSON LOG — full nested structure
    # ------------------------------------------------------------------
    json_path = OUTPUT_FOLDER / f"{log_stem}.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(run_log, f, indent=2, ensure_ascii=False)
    print(f"Defensibility log (JSON) -> {json_path}")

    # ------------------------------------------------------------------
    # CSV LOG — flat, sectioned format for human review
    # ------------------------------------------------------------------
    csv_path = OUTPUT_FOLDER / f"{log_stem}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        writer.writerow(["SECTION", "FIELD", "VALUE"])
        writer.writerow(["RUN_HEADER", "run_id", run_log["run_id"]])
        writer.writerow(["RUN_HEADER", "run_timestamp", run_log["run_timestamp"]])
        writer.writerow(["RUN_HEADER", "script_version", run_log["script_version"]])
        writer.writerow(["RUN_HEADER", "root_folder", run_log["root_folder"]])
        writer.writerow(["RUN_HEADER", "output_folder", run_log["output_folder"]])
        writer.writerow(["RUN_HEADER", "status", run_log["status"]])
        writer.writerow(["RUN_HEADER", "status_detail", run_log["status_detail"]])
        writer.writerow([])

        writer.writerow(["SUMMARY", "FIELD", "VALUE"])
        for k, v in run_log["summary"].items():
            writer.writerow(["SUMMARY", k, v])
        writer.writerow([])

        writer.writerow(["INPUTS", "file_name", "file_path",
                         "file_size_bytes", "matching_csv_files", "rows_read"])
        for inp in run_log["inputs"]:
            writer.writerow([
                "INPUTS",
                inp["file_name"],
                inp["file_path"],
                inp["file_size_bytes"] if inp["file_size_bytes"] is not None else "N/A",
                "; ".join(inp["matching_csv_files"]) if inp["matching_csv_files"] else "None",
                inp["rows_read"],
            ])
        if not run_log["inputs"]:
            writer.writerow(["INPUTS", "No input files processed.", "", "", "", ""])
        writer.writerow([])

        writer.writerow(["OUTPUTS", "file_name", "group",
                         "rows_written", "file_size_bytes", "file_path"])
        for out in run_log["outputs"]:
            writer.writerow([
                "OUTPUTS",
                out["file_name"],
                out["group"],
                out["rows_written"],
                out["file_size_bytes"] if out["file_size_bytes"] is not None else "N/A",
                out["file_path"],
            ])
        if not run_log["outputs"]:
            writer.writerow(["OUTPUTS", "No output files written.", "", "", "", ""])
        writer.writerow([])

        writer.writerow(["EXCEPTIONS", "timestamp", "severity",
                         "error_type", "source_file", "detail"])
        for exc in run_log["exceptions"]:
            writer.writerow([
                "EXCEPTIONS",
                exc["timestamp"],
                exc["severity"],
                exc["error_type"],
                exc["source_file"],
                exc["detail"],
            ])
        if not run_log["exceptions"]:
            writer.writerow(["EXCEPTIONS", "", "", "No exceptions recorded.", "", ""])

    print(f"Defensibility log (CSV)  -> {csv_path}")


# -------------------------------------------------------------------
# CORE PROCESSING FUNCTIONS
# -------------------------------------------------------------------

def normalize_header(header):
    """
    Strips leading/trailing whitespace from each column name in a header row.
    Replaces None values with an empty string.
    """
    return [h.strip() if h is not None else "" for h in header]


def matches_group(filename: str):
    """
    Tests a filename against each pattern in PATTERNS.
    Returns the matching group name or None.
    """
    base = Path(filename).name
    for group, pattern in PATTERNS.items():
        if pattern.match(base):
            return group
    return None


def collect_rows_from_zip(zip_path: Path, grouped_data: dict, run_log: dict):
    """
    Opens a single zip file and extracts rows from any CSVs inside it
    that match the known naming patterns.

    For each matching CSV:
      - Reads the header row and normalises column names.
      - On the first file seen for a group, locks in the column structure.
      - Aligns subsequent files to the locked column order.
      - Appends a Source_File column recording origin.

    Records all actions in the defensibility log.
    """
    csv_files_found = []
    rows_read_this_zip = 0

    try:
        with ZipFile(zip_path, "r") as zf:
            for member in zf.namelist():
                if member.endswith("/"):
                    continue

                group = matches_group(member)
                if not group:
                    continue

                csv_files_found.append(Path(member).name)
                run_log["summary"]["csv_files_matched"] += 1
                source_file = Path(member).name

                try:
                    with zf.open(member) as f:
                        lines = (line.decode("utf-8-sig") for line in f)
                        reader = csv.DictReader(lines)

                        if reader.fieldnames is None:
                            log_exception(
                                run_log,
                                error_type="EMPTY_CSV_HEADER",
                                detail=(
                                    f"CSV file '{member}' inside '{zip_path.name}' "
                                    f"has no header row. File skipped."
                                ),
                                source_file=source_file,
                                severity="WARNING",
                            )
                            continue

                        current_header = normalize_header(reader.fieldnames)

                        if grouped_data[group]["fieldnames"] is None:
                            grouped_data[group]["fieldnames"] = current_header + ["Source_File"]
                            print(f"Using header from first {group} file: {member}")
                        else:
                            ref_cols = set(grouped_data[group]["fieldnames"]) - {"Source_File"}
                            cur_cols = set(current_header)
                            if ref_cols != cur_cols:
                                extra = cur_cols - ref_cols
                                missing = ref_cols - cur_cols
                                log_exception(
                                    run_log,
                                    error_type="SCHEMA_MISMATCH",
                                    detail=(
                                        f"CSV file '{member}' inside '{zip_path.name}' "
                                        f"has a different column set from the first "
                                        f"{group} file. "
                                        f"Extra columns (discarded): {sorted(extra)}. "
                                        f"Missing columns (filled empty): {sorted(missing)}."
                                    ),
                                    source_file=source_file,
                                    severity="WARNING",
                                )

                        first_header = grouped_data[group]["fieldnames"]
                        rows_this_file = 0

                        for row in reader:
                            cleaned_row = {}
                            for key, value in row.items():
                                clean_key = key.strip() if key else ""
                                cleaned_row[clean_key] = value

                            ordered_row = {}
                            for col in first_header[:-1]:  # Exclude Source_File
                                ordered_row[col] = cleaned_row.get(col, "")

                            ordered_row["Source_File"] = source_file
                            grouped_data[group]["rows"].append(ordered_row)
                            rows_this_file += 1

                        rows_read_this_zip += rows_this_file
                        run_log["summary"]["total_rows_read"] += rows_this_file

                except Exception as e:
                    log_exception(
                        run_log,
                        error_type="CSV_READ_ERROR",
                        detail=(
                            f"Failed to read CSV '{member}' inside '{zip_path.name}'. "
                            f"Error: {e}. Traceback: {traceback.format_exc(limit=3)}"
                        ),
                        source_file=source_file,
                        severity="ERROR",
                    )

    except BadZipFile:
        log_exception(
            run_log,
            error_type="BAD_ZIP_FILE",
            detail=(
                f"Zip file '{zip_path.name}' is corrupt or invalid and could not "
                f"be opened. File skipped. All CSVs inside this zip are excluded "
                f"from the combined output."
            ),
            source_file=zip_path.name,
            severity="ERROR",
        )
        run_log["summary"]["zip_files_skipped"] += 1
        return

    except Exception as e:
        log_exception(
            run_log,
            error_type="ZIP_READ_ERROR",
            detail=(
                f"Unexpected error processing zip file '{zip_path.name}'. "
                f"Error: {e}. Traceback: {traceback.format_exc(limit=3)}"
            ),
            source_file=zip_path.name,
            severity="ERROR",
        )
        run_log["summary"]["zip_files_skipped"] += 1
        return

    log_input(run_log, zip_path, csv_files_found, rows_read_this_zip)
    run_log["summary"]["zip_files_processed"] += 1


def deduplicate_rows(grouped_data: dict, run_log: dict):
    """
    Removes duplicate rows from each group's accumulated row list.

    Deduplication compares all columns except Source_File — so if the same
    data row appears in multiple zip files, only the first occurrence is kept.
    """
    for group, data in grouped_data.items():
        if not data["rows"]:
            continue

        deduped_rows = []
        seen = set()
        compare_columns = [col for col in data["fieldnames"] if col != "Source_File"]

        for row in data["rows"]:
            key = tuple((row.get(col, "") or "").strip() for col in compare_columns)
            if key not in seen:
                seen.add(key)
                deduped_rows.append(row)

        removed_count = len(data["rows"]) - len(deduped_rows)
        data["rows"] = deduped_rows
        run_log["summary"]["total_duplicate_rows_removed"] += removed_count
        print(f"{group}: removed {removed_count} duplicate rows")

        if removed_count > 0:
            log_exception(
                run_log,
                error_type="DUPLICATE_ROWS_REMOVED",
                detail=(
                    f"Group '{group}': {removed_count} duplicate row(s) removed "
                    f"during deduplication. Surviving row retains the Source_File "
                    f"value from the first zip file processed containing that row."
                ),
                source_file="",
                severity="WARNING",
            )


def write_output(grouped_data: dict, run_log: dict, run_date: str):
    """
    Writes one combined Excel (.xlsx) file per group to OUTPUT_FOLDER.

    Each file contains:
      - A styled header row (bold white text on dark blue background).
      - All deduplicated data rows.
      - Auto-sized column widths based on header label length.

    Skips groups where no rows were collected.
    Records each written file in the defensibility log.
    """
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    for group, data in grouped_data.items():
        if not data["rows"]:
            print(f"No files found for {group}")
            continue

        output_path = OUTPUT_FOLDER / f"{OUTPUT_FILE_BASES[group]}_{run_date}.xlsx"

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = group

            # Write header row
            ws.append(data["fieldnames"])

            # Write data rows
            for row in data["rows"]:
                ws.append([row.get(col, "") for col in data["fieldnames"]])

            # Apply header styling and column widths
            apply_header_style(ws)

            # Freeze the header row so it stays visible when scrolling
            ws.freeze_panes = "A2"

            wb.save(output_path)

            log_output(run_log, output_path, group, len(data["rows"]))
            print(f"{group} combined -> {output_path}")

        except Exception as e:
            log_exception(
                run_log,
                error_type="OUTPUT_WRITE_ERROR",
                detail=(
                    f"Failed to write combined output for group '{group}' "
                    f"to '{output_path}'. Error: {e}. "
                    f"Traceback: {traceback.format_exc(limit=3)}"
                ),
                source_file=OUTPUT_FILES[group],
                severity="ERROR",
            )


def main():
    """
    Entry point. Orchestrates the full pipeline:
      1. Initialises the defensibility log and run timestamp.
      2. Validates that the root input folder exists.
      3. Initialises the grouped_data structure.
      4. Discovers all zip files recursively under ROOT_FOLDER.
      5. Calls collect_rows_from_zip for each zip file found.
      6. Deduplicates rows across all sources.
      7. Writes the combined Excel output files.
      8. Sets final run status and writes the defensibility log.
    """
    run_timestamp = datetime.now().strftime("%Y%m%d")
    run_log = init_run_log(run_timestamp)

    try:
        if not ROOT_FOLDER.exists():
            log_exception(
                run_log,
                error_type="ROOT_FOLDER_NOT_FOUND",
                detail=(
                    f"Root input folder does not exist: '{ROOT_FOLDER}'. "
                    f"No files were processed. Verify the ROOT_FOLDER path "
                    f"at the top of the script."
                ),
                source_file="",
                severity="CRITICAL",
            )
            run_log["status_detail"] = (
                f"Root input folder not found: '{ROOT_FOLDER}'. Run aborted."
            )
            print(f"Folder not found: {ROOT_FOLDER}")
            write_defensibility_log(run_log, run_timestamp)
            return

        grouped_data = {
            "Rename_log": {"fieldnames": None, "rows": []},
            "Rename_list": {"fieldnames": None, "rows": []},
            "Exceptions_list": {"fieldnames": None, "rows": []},
        }

        zip_files = list(ROOT_FOLDER.rglob("*.zip"))
        run_log["summary"]["zip_files_found"] = len(zip_files)

        if not zip_files:
            log_exception(
                run_log,
                error_type="NO_ZIP_FILES_FOUND",
                detail=(
                    f"No zip files were found under '{ROOT_FOLDER}' "
                    f"or any of its subfolders. No output was produced."
                ),
                source_file="",
                severity="WARNING",
            )
            run_log["status_detail"] = "No zip files found. No output produced."
            print("No zip files found.")
            write_defensibility_log(run_log, run_timestamp)
            return

        for zip_path in zip_files:
            print(f"Processing: {zip_path}")
            collect_rows_from_zip(zip_path, grouped_data, run_log)

        deduplicate_rows(grouped_data, run_log)
        write_output(grouped_data, run_log, run_timestamp)

        s = run_log["summary"]
        run_log["status_detail"] = (
            f"Processed {s['zip_files_processed']} of {s['zip_files_found']} zip file(s). "
            f"{s['total_rows_read']} rows read, "
            f"{s['total_duplicate_rows_removed']} duplicates removed, "
            f"{s['total_rows_written']} rows written across "
            f"{s['output_files_written']} output file(s). "
            f"{s['exception_count']} exception(s) recorded."
        )

    except Exception as e:
        log_exception(
            run_log,
            error_type="UNEXPECTED_ERROR",
            detail=(
                f"An unexpected error occurred during the run. "
                f"Error: {e}. Traceback: {traceback.format_exc()}"
            ),
            source_file="",
            severity="CRITICAL",
        )
        run_log["status_detail"] = f"Run failed due to an unexpected error: {e}"

    finally:
        write_defensibility_log(run_log, run_timestamp)


if __name__ == "__main__":
    main()
